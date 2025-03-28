import streamlit as st
from dotenv import load_dotenv
import os
import tempfile
import PyPDF2
import docx2txt
import openpyxl
import pandas as pd
from datetime import datetime, timedelta
from ortools.sat.python import cp_model
from fpdf import FPDF

from langchain.chat_models import ChatOpenAI
from langchain.chains.question_answering import load_qa_chain
from langchain.document_loaders import PyPDFLoader, UnstructuredWordDocumentLoader, UnstructuredExcelLoader
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain.embeddings import OpenAIEmbeddings
from langchain.vectorstores import FAISS

load_dotenv()

st.set_page_config(page_title="My Personal Assistant", layout="wide")
st.title("📁 My Personal Assistant")

if "chain" not in st.session_state:
    st.session_state.chain = None

uploaded_file = st.file_uploader("Upload a document (PDF, Word, Excel)", type=["pdf", "docx", "doc", "xlsx"])

# ────────────────────────────────────────────────
# 🔍 Document Scanner (PDF, Word, Excel)
# ────────────────────────────────────────────────
def scan_document_for_keyword(file, keyword, file_ext):
    results = []

    if file_ext == "pdf":
        reader = PyPDF2.PdfReader(file)
        for page_num, page in enumerate(reader.pages):
            try:
                text = page.extract_text()
                if keyword.lower() in text.lower():
                    results.append((f"Page {page_num+1}", text[:500]))
            except:
                continue

    elif file_ext in ["doc", "docx"]:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".docx") as tmp:
            tmp.write(file.read())
            tmp_path = tmp.name
        paragraphs = docx2txt.process(tmp_path).split("\n")
        for idx, para in enumerate(paragraphs):
            if keyword.lower() in para.lower():
                results.append((f"Paragraph {idx+1}", para.strip()[:500]))

    elif file_ext == "xlsx":
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(file.read())
            tmp_path = tmp.name
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                row_text = " ".join([str(cell) for cell in row if cell])
                if keyword.lower() in row_text.lower():
                    results.append((f"Sheet: {sheet.title}", row_text.strip()[:500]))

    return results

# ────────────────────────────────────────────────
# 🤖 Langchain AI Q&A Mode
# ────────────────────────────────────────────────
def process_uploaded_file(uploaded_file):
    with tempfile.NamedTemporaryFile(delete=False) as tmp_file:
        tmp_file.write(uploaded_file.read())
        tmp_file_path = tmp_file.name

    file_ext = uploaded_file.name.split(".")[-1].lower()

    if file_ext == "pdf":
        loader = PyPDFLoader(tmp_file_path)
    elif file_ext in ["doc", "docx"]:
        loader = UnstructuredWordDocumentLoader(tmp_file_path)
    elif file_ext == "xlsx":
        loader = UnstructuredExcelLoader(tmp_file_path)
    else:
        st.error("Unsupported file type")
        return None

    return loader.load()

@st.cache_resource
def create_chain(_docs):
    splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=100)
    texts = splitter.split_documents(_docs)

    embeddings = OpenAIEmbeddings()
    db = FAISS.from_documents(texts, embeddings)
    retriever = db.as_retriever()

    llm = ChatOpenAI(temperature=0)
    chain = load_qa_chain(llm, chain_type="stuff")

    return retriever, chain

# ────────────────────────────────────────────────
# 📅 Schedule Conflict Checker & Auto-Rescheduling
# ────────────────────────────────────────────────
def detect_schedule_conflicts_and_reschedule(file):
    df = pd.read_excel(file)
    st.write("### Schedule Preview:")
    st.dataframe(df)

    if not {'Name', 'Start Time', 'End Time'}.issubset(df.columns):
        st.error("Excel must include columns: Name, Start Time, End Time")
        return

    df['Start Time'] = pd.to_datetime(df['Start Time'])
    df['End Time'] = pd.to_datetime(df['End Time'])

    # User Input for Constraints
    min_hours_between_shifts = st.number_input("Minimum hours between shifts:", min_value=0, max_value=12, value=2)
    max_shifts_per_day = st.number_input("Maximum shifts per day:", min_value=1, max_value=5, value=3)
    
    # Constraint model setup
    model = cp_model.CpModel()
    variables = {}

    # Add variables for each shift (Start and End Time)
    for index, row in df.iterrows():
        shift_start = model.NewIntVar(0, 1000000, f"start_{index}")
        shift_end = model.NewIntVar(0, 1000000, f"end_{index}")
        variables[index] = {'start': shift_start, 'end': shift_end}

    # Add constraints for no overlap: If a person has two shifts that overlap, raise an exception
    for i, row1 in df.iterrows():
        for j, row2 in df.iterrows():
            if i >= j:
                continue
            if row1['Name'] == row2['Name']:
                model.Add(variables[i]['end'] <= variables[j]['start'])  # No overlap

                # Add minimum hours constraint between shifts
                model.Add(variables[j]['start'] - variables[i]['end'] >= min_hours_between_shifts * 60)

    # Max shifts per day constraint
    for name, group in df.groupby('Name'):
        for i, row in group.iterrows():
            model.Add(sum(variables[idx]['start'] for idx in group.index) <= max_shifts_per_day)

    # Solve the problem
    solver = cp_model.CpSolver()
    status = solver.Solve(model)

    if status == cp_model.OPTIMAL:
        st.success("Rescheduling completed successfully.")
        for index in df.index:
            df.at[index, 'Start Time'] = str(solver.Value(variables[index]['start']))
            df.at[index, 'End Time'] = str(solver.Value(variables[index]['end']))
        st.write("### Resolved Schedule:")
        st.dataframe(df)
        
        # Download buttons for Excel and PDF
        download_excel(df)
        download_pdf(df)
        
    else:
        st.error("No solution found for conflicts.")

# ────────────────────────────────────────────────
# Download Options
# ────────────────────────────────────────────────
def download_excel(df):
    # Save the resolved schedule to a new Excel file
    file_path = "/mnt/data/optimized_schedule.xlsx"
    df.to_excel(file_path, index=False)
    st.download_button("Download Optimized Schedule (Excel)", file_path)

def download_pdf(df):
    # Convert the DataFrame to a PDF file
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    
    pdf.cell(200, 10, txt="Optimized Schedule", ln=True, align="C")
    
    # Add table headers
    pdf.cell(40, 10, "Name", 1)
    pdf.cell(50, 10, "Start Time", 1)
    pdf.cell(50, 10, "End Time", 1)
    pdf.ln()
    
    # Add rows
    for index, row in df.iterrows():
        pdf.cell(40, 10, row['Name'], 1)
        pdf.cell(50, 10, str(row['Start Time']), 1)
        pdf.cell(50, 10, str(row['End Time']), 1)
        pdf.ln()
    
    # Save PDF
    file_path = "/mnt/data/optimized_schedule.pdf"
    pdf.output(file_path)
    st.download_button("Download Optimized Schedule (PDF)", file_path)

# ────────────────────────────────────────────────
# 📄 PDF Page Extraction Based on Route Name
# ────────────────────────────────────────────────
def extract_pdf_pages_by_route(file, route_name):
    reader = PyPDF2.PdfReader(file)
    writer = PyPDF2.PdfWriter()

    found_pages = []
    
    # Iterate over all pages in the PDF
    for page_num, page in enumerate(reader.pages):
        try:
            text = page.extract_text()
            if text and route_name.lower() in text.lower():  # Check if route_name is in page text
                found_pages.append(page_num)
                writer.add_page(reader.pages[page_num])  # Add page to writer
        except Exception as e:
            continue

    if found_pages:
        # Save the extracted pages to a new PDF file
        output_pdf = "/mnt/data/extracted_pages_by_route.pdf"
        with open(output_pdf, "wb") as output_file:
            writer.write(output_file)
        return output_pdf, found_pages
    else:
        return None, None

# ────────────────────────────────────────────────
# UI Mode Logic: Allow users to choose route and extract pages
# ────────────────────────────────────────────────
mode = st.radio("Choose mode:", ["🔍 Document Scanner", "🤖 AI Chat", "📅 Schedule Conflict Checker", "📄 Extract PDF Pages by Route"])

if uploaded_file:
    file_ext = uploaded_file.name.split(".")[-1].lower()

    if mode == "📄 Extract PDF Pages by Route":
        if file_ext == "pdf":
            # Ask user for the route name to search
            route_name = st.text_input("Enter the route name or keyword to search for:")
            if route_name:
                with st.spinner(f"Searching for '{route_name}' in PDF..."):
                    extracted_pdf, found_pages = extract_pdf_pages_by_route(uploaded_file, route_name)
                    if extracted_pdf:
                        st.success(f"Found {len(found_pages)} page(s) with '{route_name}'!")
                        st.write(f"Pages Found: {', '.join(map(str, [page + 1 for page in found_pages]))}")  # Display page numbers (1-indexed)
                        st.download_button("Download Extracted PDF", extracted_pdf)
                    else:
                        st.warning(f"No pages found with route name '{route_name}'.")
        else:
            st.error("Please upload a valid PDF file to extract pages.")
    
    elif mode == "🔍 Document Scanner":
        search_term = st.text_input("Enter a name or keyword to search for:")
        if search_term:
            with st.spinner("Scanning document..."):
                results = scan_document_for_keyword(uploaded_file, search_term, file_ext)
                if results:
                    st.success(f"Found {len(results)} match(es) for '{search_term}':")
                    for label, snippet in results:
                        with st.expander(f"🔎 {label}"):
                            st.write(snippet)
                else:
                    st.warning("No matches found.")

    elif mode == "🤖 AI Chat":
        if "last_uploaded_filename" not in st.session_state or uploaded_file.name != st.session_state.last_uploaded_filename:
            with st.spinner("Processing document for AI..."):
                docs = process_uploaded_file(uploaded_file)
                if docs:
                    retriever, chain = create_chain(docs)
                    st.session_state.chain = chain
                    st.session_state.retriever = retriever
                    st.session_state.last_uploaded_filename = uploaded_file.name
                    st.success("Document processed! Ask your question below:")

        if st.session_state.chain:
            user_question = st.text_input("Ask a question about your document:")
            if user_question:
                relevant_docs = st.session_state.retriever.get_relevant_documents(user_question)
                answer = st.session_state.chain.run(input_documents=relevant_docs, question=user_question)
                st.write("🤖", answer)

    elif mode == "📅 Schedule Conflict Checker":
        if file_ext != "xlsx":
            st.error("Please upload a valid Excel (.xlsx) schedule file.")
        else:
            detect_schedule_conflicts_and_reschedule(uploaded_file)
else:
    st.info("📤 Upload a document to get started.")

